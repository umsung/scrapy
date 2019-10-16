# -*- coding: utf-8 -*-

# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://doc.scrapy.org/en/latest/topics/item-pipeline.html
import csv
import json
import os
import re
import time
from hashlib import md5
import numpy as np
import pymongo
import pymysql
import requests
from scrapy.exceptions import DropItem
import pandas as pd
import redis
from openpyxl import Workbook
import xlwt
import numpy as np
from quotetutorail.items import WeiboItem


class QuotetPipeline(object):
    '''
#保存为CSV格式
    '''
    def __init__(self):
        self.file = open('baidu1.csv', 'a', encoding='utf-8', newline='')

    def process_item(self, item, spider):
        a = csv.writer(self.file)
        a.writerow(item.values())
        return item

#
# class ImgSavePipeline(object):
#     '''
# #创建文件夹，下载保存图片
#     '''
#     def __init__(self):
#         try:
#             os.mkdir('图片')
#         except FileExistsError:
#             pass
#
#     def process_item(self, item, spider):
#
#         resp = requests.get(item['img_href'])
#         img_html = resp.content
#         if resp.status_code == 200:
#             with open('./test/{}.jpg'.format(item['title']), 'wb') as f:
#                 f.write(img_html)
#                 return item
#         else:
#                 return DropItem(item)


class ExcelPipline(object):
    def __init__(self):
        self.wb = xlwt.Workbook(encoding='utf-8')
        self.ws = self.wb.add_sheet('sheet1', cell_overwrite_ok=True)

    def process_item(self, item, spider):
        header = ['1', '2', '3', '4', '5']
        for i in range(len(header)):
            self.ws.write(0, i, header[i])

        x = 1
        for y in range(len(item.values())):
            self.ws.write(x, 0, item['number'])
            self.ws.write(x, 1, item['address'])
            self.ws.write(x, 2, item['title'])
            self.ws.write(x, 3, item['gps'])
            self.ws.write(x, 4, item['tengxun'])
            self.ws.write(x, 5, item['baidu'])
            x = x+1
        self.wb.save('E:\ke.xlsx')
        return item


class TestPanadsPipline(object):
    def __init__(self):
        # 创建excel，填写表头
        self.data = []

    def process_item(self, item, spider):

        self.data.append(list(item.values()))
        df = pd.DataFrame(self.data, columns=['snapshot_url', 'real_url', '', '', '', '', ''])

        # df = pd.DataFrame(self.data, columns=['名称', '地址', '电话', 'gps', 'tengxun', 'baidu'])
        # df = df[df['电话'].str.contains('视觉.*?艺术')]
        # 协会关键词
        # if df['名称'].str.contains('新媒体.*?协会'):
        #     df = df[df['名称'].str.contains('新媒体.*?协会')]
        #     df = df[df['名称'].str.contains('新媒体.*?协会')]
        #
        # if df['名称'].str.contains('摄影家.*?协会'):
        #     df = df[df['名称'].str.contains('摄影家.*?协会')]
        #
        # if df['名称'].str.contains('动画.*?协会'):
        #     df = df[df['名称'].str.contains('动画.*?协会')]
        #
        # if df['名称'].str.contains('动漫.*?协会'):
        #     df = df[df['名称'].str.contains('动漫.*?协会')]
        #
        # if df['名称'].str.contains('游戏.*?协会'):
        #     df = df[df['名称'].str.contains('游戏.*?协会')]

        df.to_excel('D:\\交易所.xlsx', encoding='utf-8', index=0)

        return item



class EntePipline(object):
    def __init__(self):
        # 创建excel，填写表头
        self.wb = Workbook()   # 创建表对象
        self.ws = self.wb.active  # 创建一个工作表 如果想改名字，可以直接给title属性赋值 ws.title = "New Shit"
        # self.ws = self.wb[self.wb.sheetnames[0]]
        self.ws.append(['名称', '地址','电话'])  # 设置表头


    def process_item(self, item, spider):

        self.ws.append(list(item.values()))
        self.ws.column_dimensions['A'].auto_size = True
        self.wb.save('D:\\北京-动漫公司04.xlsx')
        return item
        #
        # for i in range(len(item["name"])):
        #     line = [(item['name'][i]), (item['address'][i]), (item['tel'][i]).replace('电话：', "")]
        #     self.ws.append(line)
        #     self.wb.save('D:\\北京-动漫公司01.xlsx')
            # 读取Excel中Sheet1中的数据
            # data = pd.read_excel('F:\\北京-动漫公司01.xlsx')

            # 查看读取数据内容
            # print(data)
            #
            # # 查看是否有重复行
            # re_row = data.duplicated()
            # print(re_row)
            #
            # # 查看去除重复行的数据
            # no_re_row = data.drop_duplicates()
            # print(no_re_row)
            #
            # # 查看基于[名称]列去除重复行的数据
            # wp = data.drop_duplicates(['名称', "地址", "电话"])
            # # print(wp)
            #
            # # 将去除重复行的数据输出到excel表中
            # wp.to_excel("F:\\北京-动漫公司02.xlsx",index = False)

            # return item



class EntPipline(object):
    def __init__(self):
        # 创建excel，填写表头
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.append(['名称', '地址','电话'])  # 设置表头

    def process_item(self,item,spider):

        for i in range(len(item["name"])):
            line = [(item['name'][i]), (item['address'][i]), (item['tel'][i]).replace('电话：', "")]
            self.ws.append(line)
            self.wb.save("F:\\"+spider.city+'-'+spider.keyword+".xlsx")
            # 读取Excel中Sheet1中的数据
            data = pd.read_excel("F:\\"+spider.city+'-'+spider.keyword+".xlsx")

            # 查看基于[名称]列去除重复行的数据
            data.drop_duplicates(['名称', "地址", "电话"])

################################过滤非关键词#############################################
            #企业关键词
            if spider.keyword == '文化产业有限公司':
                bool = data['名称'].str.contains('文化产业.*?公司')
                filter_data = data[bool]
            if spider.keyword == "视觉艺术交流有限公司":
                bool = data['名称'].str.contains('视觉.*?艺术.*?公司')
                filter_data = data[bool]
            if spider.keyword == '动漫设计股份有限公司':
                bool = data['名称'].str.contains('动漫.*?公司')
                filter_data = data[bool]
            # 将去除重复行的数据输出到excel表中
            filter_data.to_excel("F:\\"+spider.city+'-'+spider.keyword+".xlsx",index = False)

            return item


# 保存成json格式
class JsonWriterPipeline(object):

    def __init__(self):
        self.file = open('items.txt', 'a', encoding='utf-8')

    def process_item(self, item, spider):
        line = json.dumps(dict(item), ensure_ascii=False) + "\n"
        self.file.write(line)
        return item

class WeiboPipeline(object):
    # time.strftime('%Y{y}%m{m}%d{d} %H{h}%M{f}%S{s}').format(y='年', m='月', d='日', h='时', f='分', s='秒')
    def parse_time(self, datetime):
        if re.match('\d+月\d+日', datetime):
            datetime = time.strftime('%Y年', time.localtime()) + datetime
        if re.match('\d+分钟前', datetime):
            minute = re.match('(\d+)', datetime).group(1)
            datetime = time.strftime('%Y年%m月%d日 %H:%M', time.localtime(time.time() - float(minute) * 60))
        if re.match('今天.*', datetime):
            datetime = re.match('今天(.*)', datetime).group(1).strip()
            datetime = time.strftime('%Y年%m月%d日', time.localtime()) + ' ' + datetime
        return datetime

    def process_item(self, item, spider):
        if isinstance(item, WeiboItem):
            if item.get('content'):
                item['content'] = item['content'].lstrip(':').strip()
            if item.get('posted_at'):
                item['posted_at'] = item['posted_at'].strip()
                item['posted_at'] = self.parse_time(item.get('posted_at'))
        return item


redis_db = redis.Redis(host='localhost', port=6379, db=1)  # 连接本地redis，db数据库默认连接到0号库，写的是索引值
redis_data_dict = 'test'   # redis字典名称, 相当于数据表名

#插入到Mysql数据库
class  MysqlPipeline(object):
    def __init__(self):
        #连接数据库
        self.conn = pymysql.connect(host='localhost', port=3306, user='root', passwd='root', db='testdb', use_unicode=True, charset='utf8mb4')
        #创建cursor对象执行sql语句
        self.cursor = self.conn.cursor()
        # self.cursor.execute('drop table if exists dTable')
        # 创建数据表
        # create_sql = '''create table if not exists dTable(
        #                 title varchar(300),
        #                 img_href varchar(5000)
        # )Default charset=utf8mb4
        # '''
        # self.cursor.execute(create_sql)
        # self.cursor.execute('alter table dTable convert to character set utf8mb4')

        redis_db.flushdb()  # 清空当前数据库中的所有 key，为了后面将mysql数据库中的数据全部保存进去
        # print(redis_db)
        if redis_db.hlen(redis_data_dict) == 0:  # 判断redis数据库中的key，若不存在就读取mysql数据并临时保存在redis中
            sql = 'select img_href from dTable'  # 查询表中的现有数据
            df = pd.read_sql(sql, self.conn)  # 读取mysql中的数据
            print(df)
            for url in df['img_href'].get_values():
                redis_db.hset(redis_data_dict, url, 0)  # 把每个url写入field中，value值随便设，我设置的0  key field value 三者的关系

    def process_item(self, item, spider):
        """
        比较爬取的数据在数据库中是否存在，不存在则插入数据库
        :param item: 爬取到的数据
        :param spider: /
        """
        if redis_db.hexists(redis_data_dict, item['url']):  # 比较的是redis_data_dict里面的field

            print("数据库已经存在该条数据，不再继续追加")
            raise DropItem("Duplicate item found: %s" % item)
        else:
            self.do_insert(item)

    def do_insert(self, item):

        #插入数据
        insert_sql = '''insert into dTable(title, img_href, url) values(%s, %s, %s)'''

        self.cursor.execute(insert_sql, (item['title'], item['img_href'], item['url']))
        self.conn.commit()
        # return item

    def close_spider(self, spider):
        self.conn.close()


class MongoPipeline(object):
    collection_name = 'scrapy_items'
    def __init__(self, mongo_uri, mongo_db):
        self.mongo_uri = mongo_uri
        self.mongo_db = mongo_db

    @classmethod
    def from_crawler(cls, crawler):
        return cls(
            mongo_uri=crawler.settings.get('MONGO_URI'),
            mongo_db=crawler.settings.get('MONGO_DATABASE', 'items')
        )

    def open_spider(self, spider):
        # 用MongoClient指定连接数据库的地址self.mongo_uri,创建MongoClient对象client
        self.client = pymongo.MongoClient(self.mongo_uri)
        # 在数据库url中指定要创建的数据库self.mongo_db
        self.db = self.client[self.mongo_db]

    def close_spider(self, spider):
        self.client.close()

    def process_item(self, item, spider):
        self.db[self.collection_name].insert(dict(item))
        return item

